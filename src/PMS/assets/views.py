#!/usr/bin/python
# coding=utf-8
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from assets.forms import AssetModelForm, AssetSearchForm
from assets.models import Asset, AssetArea, AssetCategory, AssetStatus, AssetType, Brand, Doc_attachment, Label_attachment, Pic_attachment, Series, Location, Unit
import openpyxl
from django.http import JsonResponse
import os
from PMS.settings.base import ASSET_BTW_FILE, EXE_FILE, NON_ASSET_BTW_FILE, PRINTER, BASE_DIR
import csv
from datetime import datetime
from django.shortcuts import render, redirect
from django.urls import reverse
from django.conf import settings
from django.http import HttpResponse, HttpResponseNotFound, Http404
from django.core.paginator import Paginator
import xlwt
from django.db.models import Q

def print_cmd(EXCEL_FILE, BTW_FILE):
    CMD = """{EXE_FILE} /AF=\"{BTW_FILE}\" /D=\"{EXCEL_FILE}\" /PRN=\"{PRINTER}\" /P/X""".format(EXE_FILE=EXE_FILE, BTW_FILE=BTW_FILE, EXCEL_FILE=EXCEL_FILE, PRINTER=PRINTER)
    print(CMD)
    os.system(CMD)

@login_required
def index(request):
    return render(request, 'assets/index.html', locals())

def get_main_url(request):
    return reverse('assets_main')

#刪除
def delete(request, pk):
    if request.method == "GET":
        asset = Asset.objects.get(id=pk)
        asset.delete()
        return redirect(get_main_url(request))
    return render(request, 'assets/main.html', locals())

#Excel
def export_assets_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="assets.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Users')
    ws.col(0).width = 256 *20
    ws.col(1).width = 256 *20
    ws.col(2).width = 256 *20
    ws.col(3).width = 256 *20
    ws.col(4).width = 256 *20
    ws.col(5).width = 256 *20
    ws.col(6).width = 256 *20
    ws.col(7).width = 256 *20
    ws.col(8).width = 256 *20
    ws.col(9).width = 256 *20
    ws.col(10).width = 256 *20
    ws.col(11).width = 256 *20
    ws.col(12).width = 256 *20
    ws.col(13).width = 256 *20
    ws.col(14).width = 256 *20

    # Sheet header, first row
    row_num = 0

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['資產編號', '資產類別', '資產種類', '品牌', '型號', '地區', '負責單位', '保管單位', '保管人姓名', '放置地點', '放置地點描述', '採購日期', '採購金額', '狀態', '描述']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()

    assets = get_assets_queryset(request)

    for data in assets:
        row_num += 1
        if data.keeper_unit:
            _keeper_unit = data.keeper_unit.unit_name
        else:
            _keeper_unit = ""

        ws.write(row_num, 0, data.asset_no, font_style)
        ws.write(row_num, 1, data.category.category_name, font_style)
        ws.write(row_num, 2, data.type.type_name, font_style)
        ws.write(row_num, 3, data.brand.brand_name, font_style)
        ws.write(row_num, 4, data.model, font_style)
        ws.write(row_num, 5, data.area.area_name, font_style)
        ws.write(row_num, 6, data.owner_unit.unit_name, font_style)
        ws.write(row_num, 7, _keeper_unit, font_style)
        ws.write(row_num, 8, data.keeper_name, font_style)
        ws.write(row_num, 9, data.location.location_name, font_style)
        ws.write(row_num, 10, data.location_desc, font_style)
        ws.write(row_num, 11, data.pur_date, font_style)
        ws.write(row_num, 12, data.pur_price, font_style)
        ws.write(row_num, 13, data.status.status_name, font_style)
        ws.write(row_num, 14, data.desc, font_style)

    wb.save(response)
    return response


#分頁查詢
def get_assets_queryset(request):
    assets = Asset.objects.all()
    if 'asset_no' in request.session:
        _asset_no = request.session['asset_no']
        assets = assets.filter(asset_no__icontains=_asset_no)

    if 'status' in request.session:
        _status = request.session['status']
        assets = assets.filter(status=_status)

    if 'category' in request.session:
        _category = request.session['category']
        assets = assets.filter(category=_category)

    if 'type' in request.session:
        _type = request.session['type']
        assets = assets.filter(type=_type)

    if 'area' in request.session:
        _area = request.session['area']
        assets = assets.filter(area=_area)

    if 'location' in request.session:
        _location = request.session['location']
        assets = assets.filter(location=_location)

    if 'location_desc' in request.session:
        _location_desc = request.session['location_desc']
        assets = assets.filter(location_desc__icontains=_location)

    if 'scrap' in request.session:
        _scrap = request.session['scrap']
        assets = assets.exclude(status=AssetStatus.objects.get(status_name="已報廢"))

    if 'keeper_unit' in request.session:
        _keeper_unit = request.session['keeper_unit']
        assets = assets.filter(keeper_unit=_keeper_unit)

    if 'keeper_name' in request.session:
        _keeper_name = request.session['keeper_name']
        assets = assets.filter(keeper_name__icontains=_keeper_name)

    if 'desc' in request.session:
        _desc = request.session['desc']
        assets = assets.filter(desc__icontains=_desc)

    results = assets.order_by('-type')
    return results


#查詢
def search(request):
    page_number = 1
    _asset_no = ""
    _status = ""
    _category = ""
    _type = ""
    _brand = ""
    _area = ""
    _location = ""
    _location_desc = ""
    _scrap = ""
    _keeper_unit = ""
    _keeper_name = ""
    _desc = ""
    _condition1 = ""
    _condition2 = ""
    _condition3 = ""
    _condition4 = ""
    _condition5 = ""

    assets = Asset.objects.all()
    if request.method == "POST":
        _asset_no = request.POST.get('asset_no')
        _status = request.POST.get('status')
        _category = request.POST.get('category')
        _type = request.POST.get('type')
        _brand = request.POST.get('brand')
        _area = request.POST.get('area')
        _location = request.POST.get('location')
        _location_desc = request.POST.get('location_desc')
        _scrap = request.POST.get('scrap')
        _keeper_unit = request.POST.get('keeper_unit')
        _keeper_name = request.POST.get('keeper_name')
        _desc = request.POST.get('desc')
        _condition1 = request.POST.get('condition1')
        _condition2 = request.POST.get('condition2')
        _condition3 = request.POST.get('condition3')
        _condition4 = request.POST.get('condition4')
        _condition5 = request.POST.get('condition5')

    if request.method == "GET":
        page_number = request.GET.get('page')
        if 'asset_no' in request.session:
            _asset_no = request.session['asset_no']

        if 'status' in request.session:
            _status = request.session['status']

        if 'category' in request.session:
            _category = request.session['category']

        if 'type' in request.session:
            _type = request.session['type']

        if 'brand' in request.session:
            _brand = request.session['brand']

        if 'area' in request.session:
            _area = request.session['area']

        if 'location' in request.session:
            _location = request.session['location']

        if 'location_desc' in request.session:
            _location_desc = request.session['location_desc']

        if 'scrap' in request.session:
            _scrap = request.session['scrap']

        if 'keeper_unit' in request.session:
            _keeper_unit = request.session['keeper_unit']

        if 'keeper_name' in request.session:
            _keeper_name = request.session['keeper_name']

        if 'desc' in request.session:
            _desc = request.session['desc']

        if 'condition1' in request.session:
            _condition1 = request.session['condition1']

        if 'condition2' in request.session:
            _condition2 = request.session['condition2']

        if 'condition3' in request.session:
            _condition3 = request.session['condition3']

        if 'condition4' in request.session:
            _condition4 = request.session['condition4']

        if 'condition5' in request.session:
            _condition5 = request.session['condition5']


    if _asset_no:
        request.session['asset_no'] = _asset_no
        assets = assets.filter(asset_no__icontains=_asset_no)

    if _status:
        request.session['status'] = _status
        assets = assets.filter(status=_status)

    if _category:
        request.session['category'] = _category
        assets = assets.filter(category=_category)

    if _type:
        request.session['type'] = _type
        assets = assets.filter(type=_type)

    if _brand:
        request.session['brand'] = _brand
        assets = assets.filter(brand=_brand)

    if _area:
        request.session['area'] = _area
        assets = assets.filter(area=_area)

    if _location:
        request.session['location'] = _location
        assets = assets.filter(location=_location)

    if _location_desc:
        request.session['location_desc'] = _location_desc
        assets = assets.filter(location_desc__icontains=_location)

    if _keeper_unit:
        request.session['keeper_unit'] = _keeper_unit
        assets = assets.filter(keeper_unit=_keeper_unit)

    if _keeper_name:
        request.session['keeper_name'] = _keeper_name
        assets = assets.filter(keeper_name__icontains=_keeper_name)

    if _desc:
        request.session['desc'] = _desc
        assets = assets.filter(Q(desc__icontains=_desc) | Q(comment__icontains=_desc))

    if not _scrap:
        request.session['scrap'] = _scrap
        assets = assets.exclude(status=AssetStatus.objects.get(status_name="已報廢"))

    _orderby = []
    if _condition1:
        request.session['condition1'] = _condition1
        _orderby.append(_condition1)

    if _condition2:
        request.session['condition2'] = _condition1
        _orderby.append(_condition2)

    if _condition3:
        request.session['condition3'] = _condition1
        _orderby.append(_condition3)

    if _condition4:
        request.session['condition4'] = _condition1
        _orderby.append(_condition4)

    if _condition5:
        request.session['condition5'] = _condition1
        _orderby.append(_condition5)

    if _orderby:
        assets = assets.order_by(*_orderby)

    results = list(assets)
    page_obj = Paginator(results, 200)
    row_count = assets.count()

    if page_number:
        page_results = page_obj.page(page_number)
    else:
        page_results = page_obj.page(1)

    return render(request, 'assets/search.html', locals())

#新增
def create(request):
    if request.method == "POST":
        form = AssetModelForm(request.POST)
        if form.is_valid():
            _category = form.cleaned_data.get('category')
            _type = form.cleaned_data.get('type')
            _location = form.cleaned_data.get('location')
            _auto_encode = form.cleaned_data.get('auto_encode')

            tmp_form = form.save(commit=False)
            if _auto_encode:
                tmp_form.asset_no = get_series_number(_category, _type, _location)
            tmp_form.create_by = request.user
            tmp_form.update_by = request.user
            form.save()

            if request.FILES.get('pic1'):
                request_file = Pic_attachment(
                    files=request.FILES['pic1'])
                request_file.description = request.POST['pic_desc1']
                request_file.asset = tmp_form
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('doc1'):
                request_file = Doc_attachment(
                    files=request.FILES['doc1'])
                request_file.description = request.POST['doc_desc1']
                request_file.asset = tmp_form
                request_file.create_by = request.user
                request_file.save()

            return redirect(tmp_form.get_absolute_url())
    else:
        form = AssetModelForm()

    return render(request, 'assets/edit.html', locals())

#修改
def update(request, pk):
    mode = "UPDATE"
    asset = Asset.objects.get(id=pk)
    if request.method == "POST":
        form = AssetModelForm(request.POST, instance=asset)
        if form.is_valid():
            tmp_form = form.save(commit=False)
            tmp_form.save()

            if request.FILES.get('pic1'):
                request_file = Pic_attachment(
                    files=request.FILES['pic1'])
                request_file.description = request.POST['pic_desc1']
                request_file.asset = tmp_form
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('doc1'):
                request_file = Doc_attachment(
                    files=request.FILES['doc1'])
                request_file.description = request.POST['doc_desc1']
                request_file.asset = tmp_form
                request_file.create_by = request.user
                request_file.save()

            return redirect(tmp_form.get_absolute_url())
    else:
        form = AssetModelForm(instance=asset)
        pics = Pic_attachment.objects.filter(asset=asset).all()
        docs = Doc_attachment.objects.filter(asset=asset).all()

    return render(request, 'assets/edit.html', locals())

#滾序號
def get_series_number(asset_category, asset_type, asset_location):
    type_code = asset_type.type_code  # 縮寫
    type_name = asset_type.type_name.upper()
    asset_category_id = str(asset_category.id)
    asset_category_name = str(asset_category.category_name)
    asset_location_id = str(asset_location.id)
    asset_location_name = str(asset_location.location_name)
    asset_type_id = str(asset_type.id)
    asset_type_name = str(asset_type.type_name)

    _key = ""
    _key_name = ""
    if asset_category_id == "1":  # 資訊設備
        if type_name == "NOTEBOOK":
            series_code = "TWKHHN"
        elif type_name == "COMPUTER":
            series_code = "TWKHHD"
        elif type_name == "MOBILE":
            series_code = "TWKHHM"
        else:
            series_code = "IT-" + type_code + "-"
        _key = asset_category_id.zfill(3) + asset_type_id.zfill(3)
        _key_name = asset_category_name + "_" + asset_type_name
    elif asset_category_id == "2":  # 辦公設備
        loc_obj = Location.objects.filter(id=asset_location_id)
        location_code = loc_obj[0].location_code
        series_code = "A-" + location_code + "-" + type_code + "-"
        _key = asset_category_id.zfill(3) + asset_location_id.zfill(3) + asset_type_id.zfill(3)
        _key_name = asset_category_name + "_" + asset_location_name + "_" + asset_type_name

    # 滾序號
    obj = Series.objects.filter(key=_key)
    if obj:
        _series = obj[0].series + 1
        obj.update(series=_series, desc=_key_name)
    else:
        _series = 1
        series = Series.objects.create(key=_key, series=1, desc=_key_name)

    if asset_category_id == "1":  # 資訊設備
        if type_name == "NOTEBOOK" or type_name == "COMPUTER":
            series_format = str(_series).zfill(4)
        else:
            series_format = str(_series).zfill(5)
    elif asset_category_id == "2":  # 辦公設備
        series_format = str(_series).zfill(3)

    series_code = series_code + series_format

    return series_code

#明細
def detail(request, pk):
    try:
        asset = Asset.objects.get(pk=pk)
        pics = Pic_attachment.objects.filter(asset=asset)
        docs = Doc_attachment.objects.filter(asset=asset)
    except Asset.DoesNotExist:
        raise Http404('Book does not exist')


    return render(request, 'assets/detail.html', locals())

#回主頁
def main(request):
    """清除session"""
    if 'asset_no' in request.session:
        del request.session['asset_no']
    if 'status' in request.session:
        del request.session['status']
    if 'category' in request.session:
        del request.session['category']
    if 'type' in request.session:
        del request.session['type']
    if 'area' in request.session:
        del request.session['area']
    if 'location' in request.session:
        del request.session['location']
    if 'location_desc' in request.session:
        del request.session['location_desc']
    if 'keeper_unit' in request.session:
        del request.session['keeper_unit']
    if 'keeper_name' in request.session:
        del request.session['keeper_name']
    if 'desc' in request.session:
        del request.session['desc']
    if 'scrap' in request.session:
        del request.session['scrap']
    if 'status' in request.session:
        del request.session['status']
    if 'brand' in request.session:
        del request.session['brand']

    form = AssetSearchForm()
    return render(request, 'assets/main.html', locals())

#刪除Label檔案
def delete_csv(file_name):
    try:
        os.remove(file_name)
    except OSError as e:
        print(e)
    else:
        print("File is deleted successfully")

#文字轉CSV
def TEXT2CSV(asset_number):
    now = datetime.now()
    file_name = datetime.strftime(now, '%Y%m%d %H%M%S') + ".csv"
    file_name = os.path.join(BASE_DIR, 'media', 'uploads', 'label', file_name)

    with open(file_name, 'w', newline='') as csvfile:
        fieldnames = ['NUMBER']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({'NUMBER': asset_number})
    return file_name

#印標籤專頁
def label(request):
    if request.method == 'POST':
        label_type = request.POST.get('label_type', False)

        if label_type == 'single_asset_label':
            asset_label = request.POST.get('asset_number', False)
            csv = TEXT2CSV(asset_label)
            print_cmd(csv, ASSET_BTW_FILE)
            delete_csv(csv)

        elif label_type == 'single_nonasset_label':
            asset_label = request.POST.get('non_asset', False)
            csv = TEXT2CSV(asset_label)
            print_cmd(csv, NON_ASSET_BTW_FILE)
            delete_csv(csv)
        elif label_type == 'multi_asset_label':
            excel_file = request.FILES.get('files1')
            if excel_file:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.worksheets[0]
                csv = Excel2CSV(sheet)
                print_cmd(csv, ASSET_BTW_FILE)
                delete_csv(csv)

    return render(request, 'assets/label.html', locals())

#匯入Excel
def import_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('files1')
        if excel_file:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.worksheets[0]
            for iRow in range(2, sheet.max_row+1):
                if not sheet.cell(row = iRow, column = 1).value:
                    break

                asset = Asset()
                asset.asset_no = sheet.cell(row = iRow, column = 1).value or ''
                if sheet.cell(row = iRow, column = 2).value:
                    asset.status = AssetStatus.objects.get(status_name=sheet.cell(row = iRow, column = 2).value)
                if sheet.cell(row = iRow, column = 3).value:
                    asset.category = AssetCategory.objects.get(category_name=sheet.cell(row = iRow, column = 3).value)
                if sheet.cell(row = iRow, column = 4).value:
                    asset.type = AssetType.objects.get(type_name=sheet.cell(row = iRow, column = 4).value)
                if sheet.cell(row = iRow, column = 5).value:
                    asset.brand = Brand.objects.get(brand_name=sheet.cell(row = iRow, column = 5).value)
                asset.model = sheet.cell(row = iRow, column = 6).value or ''
                asset.desc = sheet.cell(row = iRow, column = 7).value or ''
                if sheet.cell(row = iRow, column = 8).value:
                    asset.area = AssetArea.objects.get(area_name=sheet.cell(row = iRow, column = 8).value)
                asset.location = Location.objects.get(location_name=sheet.cell(row = iRow, column = 9).value)
                asset.location_desc = sheet.cell(row = iRow, column = 10).value or ''
                if sheet.cell(row = iRow, column = 11).value:
                    asset.owner_unit = Unit.objects.get(unit_name=sheet.cell(row = iRow, column = 11).value)
                if sheet.cell(row = iRow, column = 12).value:
                    asset.keeper_unit = Unit.objects.get(unit_name=sheet.cell(row = iRow, column = 12).value)
                asset.keeper_name = sheet.cell(row = iRow, column = 13).value or ''
                asset.pur_date = sheet.cell(row = iRow, column = 14).value or ''
                asset.pur_price = sheet.cell(row = iRow, column = 15).value
                asset.create_by = request.user
                asset.update_by = request.user
                asset.save()
            return redirect(get_main_url(request))

    return render(request, 'assets/import.html', locals())

#Sheet轉Table
def Sheet2Table(sheet):
    html = """<table border='1'>
                {Rows}
            </table>"""
    sRow = ""
    for iRow in range(1, sheet.max_row+1):
        sCol = ""
        for iCol in range(1, sheet.max_column+1):
            value = sheet.cell(row = iRow, column = iCol).value
            if iRow == 1:
                value = "<th>{value}</th>".format(value=value)
                sCol += value
            else:
                value = "<td>{value}</td>".format(value=value)
                sCol += value
        sCol = "<tr>" + sCol + "</tr>"
        sRow += sCol

    html = html.format(Rows=sRow)
    return html

#Sheet轉Object
def Sheet2Object(sheet):
    assets = []
    for iRow in range(2, sheet.max_row+1):
        print("目前第{iRow}行".format(iRow=iRow))
        if not sheet.cell(row = iRow, column = 1).value:
            break

        asset = Asset()
        asset.asset_no = sheet.cell(row = iRow, column = 1).value or ''
        if sheet.cell(row = iRow, column = 2).value:
            asset.status = AssetStatus.objects.get(status_name=sheet.cell(row = iRow, column = 2).value)
        if sheet.cell(row = iRow, column = 3).value:
            asset.category = AssetCategory.objects.get(category_name=sheet.cell(row = iRow, column = 3).value)
        if sheet.cell(row = iRow, column = 4).value:
            asset.type = AssetType.objects.get(type_name=sheet.cell(row = iRow, column = 4).value)
        if sheet.cell(row = iRow, column = 5).value:
            asset.brand = Brand.objects.get(brand_name=sheet.cell(row = iRow, column = 5).value)
        asset.model = sheet.cell(row = iRow, column = 6).value or ''
        asset.desc = sheet.cell(row = iRow, column = 7).value or ''
        if sheet.cell(row = iRow, column = 8).value:
            asset.area = AssetArea.objects.get(area_name=sheet.cell(row = iRow, column = 8).value)
        asset.location = Location.objects.get(location_name=sheet.cell(row = iRow, column = 9).value)
        asset.location_desc = sheet.cell(row = iRow, column = 10).value or ''
        if sheet.cell(row = iRow, column = 11).value:
            asset.owner_unit = Unit.objects.get(unit_name=sheet.cell(row = iRow, column = 11).value)
        if sheet.cell(row = iRow, column = 12).value:
            asset.keeper_unit = Unit.objects.get(unit_name=sheet.cell(row = iRow, column = 12).value)
        asset.keeper_name = sheet.cell(row = iRow, column = 13).value or ''
        asset.pur_date = sheet.cell(row = iRow, column = 14).value or ''
        asset.pur_price = sheet.cell(row = iRow, column = 15).value
        assets.append(asset)
    return assets

#Object轉Table
def Object2Table(assets, sheet):
    html = """<table border='1'>
                {Rows}
            </table>"""
    sRow = ""
    sCol = ""
    for iCol in range(1, sheet.max_column+1):
        value = sheet.cell(row = 1, column = iCol).value or '' #若為None就回傳空字串
        value = "<th>{value}</th>".format(value=value)
        sCol += value
    sCol = "<tr>" + sCol + "</tr>"
    sRow += sCol

    for asset in assets:
        sCol = ""
        sCol += "<td>{value}</td>".format(value=asset.asset_no)
        sCol += "<td>{value}</td>".format(value=asset.status)
        sCol += "<td>{value}</td>".format(value=asset.category)
        sCol += "<td>{value}</td>".format(value=asset.type)
        sCol += "<td>{value}</td>".format(value=asset.brand)
        sCol += "<td>{value}</td>".format(value=asset.model)
        sCol += "<td>{value}</td>".format(value=asset.desc)
        sCol += "<td>{value}</td>".format(value=asset.area)
        sCol += "<td>{value}</td>".format(value=asset.location)
        sCol += "<td>{value}</td>".format(value=asset.location_desc)
        sCol += "<td>{value}</td>".format(value=asset.owner_unit)
        sCol += "<td>{value}</td>".format(value=asset.keeper_unit)
        sCol += "<td>{value}</td>".format(value=asset.keeper_name)
        sCol += "<td>{value}</td>".format(value=asset.pur_date)
        sCol += "<td>{value}</td>".format(value=asset.pur_price)

        sCol = "<tr>" + sCol + "</tr>"
        sRow += sCol

    html = html.format(Rows=sRow)
    return html

#Excel轉CSV
def Excel2CSV(sheet):
    now = datetime.now()
    file_name = datetime.strftime(now, '%Y%m%d %H%M%S') + ".csv"
    file_name = os.path.join(BASE_DIR, 'media', 'uploads', 'label', file_name)

    with open(file_name, 'w', newline='') as csvfile:
        fieldnames = ['NUMBER']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for i in range(2, sheet.max_row+1):
            value = sheet.cell(row = i, column = 1).value
            if value:
                writer.writerow({'NUMBER': value})
    return file_name

#API預覽
def preview(request):
    result = ""
    if request.method == 'POST':
        excel_file = request.FILES.get('files1')
        if excel_file:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.worksheets[0]
            obj = Sheet2Object(sheet)
            result = Object2Table(obj, sheet)
    return JsonResponse(result, safe=False)

#API類別
def TypeAPI(request, category_id):
    type_data = AssetType.objects.filter(category_id = int(category_id)).values('id','type_name')
    type_list = []
    for data in type_data:
        type_list.append({'id':data['id'], 'type_name':data['type_name']})

    return JsonResponse(type_list, safe = False)

#API品牌
def BrandAPI(request, category_id):
    brand_data = Brand.objects.filter(category_id = int(category_id)).values('id','brand_name')
    brand_list = []
    for data in brand_data:
        brand_list.append({'id':data['id'], 'brand_name':data['brand_name']})

    return JsonResponse(brand_list, safe = False)

#API印標籤
def print_label(request, pk):
    asset = Asset.objects.get(pk=pk)
    now = datetime.now()
    file_name = datetime.strftime(now, '%Y%m%d %H%M%S') + ".csv"
    file_name = os.path.join(BASE_DIR, 'media', 'uploads', 'label', file_name)

    with open(file_name, 'w', newline='') as csvfile:
        fieldnames = ['NUMBER']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({'NUMBER': asset.asset_no})
    print_cmd(file_name, ASSET_BTW_FILE)
    delete_csv(file_name)

    return JsonResponse("success", safe = False)

#刪除圖片
def asset_pic_delete(request, pk):
    q = request.GET.get('q')
    if q:
        asset = Asset.objects.get(pk=q)
    obj = Pic_attachment.objects.get(pk=pk)
    if obj:
        obj.delete()
    return redirect(asset.get_absolute_url())

#刪除附件
def asset_doc_delete(request, pk):
    q = request.GET.get('q')
    if q:
        asset = Asset.objects.get(pk=q)
    obj = Doc_attachment.objects.get(pk=pk)
    if obj:
        obj.delete()
    return redirect(asset.get_absolute_url())