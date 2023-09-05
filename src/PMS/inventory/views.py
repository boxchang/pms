import json
from datetime import datetime
import openpyxl
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.db.models import Q
from django.http import JsonResponse, Http404
from django.shortcuts import render, redirect
from django.urls import reverse
from bases.utils import django_go_sql, get_invform_status_dropdown
from inventory.forms import OfficeInvForm, InvAppliedHistoryForm, OfficeItemForm, SearchForm, AttachmentForm
from inventory.models import ItemType, Item, AppliedForm, FormStatus, AppliedItem, Series, Apply_attachment, \
    ItemCategory
from users.models import CustomUser, Unit
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from datetime import datetime, timedelta


# 郵件內容
def send_template_email(subject, action, pk, address):
    if address:
        # 電子郵件內容樣板
        form = AppliedForm.objects.get(pk=pk)
        key = "{apply_date}{series}".format(apply_date=form.apply_date.replace('-', ''), series=pk)
        files = Apply_attachment.objects.filter(apply=form)
        for file in files:
            file.files.filename = file.files.name[file.files.name.rindex('/') + 1:]
        email_template = render_to_string('inventory/email_template.html', locals())

        email = EmailMessage(
            '總務用品請領單簽核通知',  # 電子郵件標題
            email_template,  # 電子郵件內容
            settings.EMAIL_HOST_USER,  # 寄件者
            [address]  # 收件者
        )
        email.fail_silently = False
        email.content_subtype = 'html'
        email.send()
        print("郵件成功寄出")

# 滾序號
def get_series_number(_key, _key_name):
    obj = Series.objects.filter(key=_key)
    if obj:
        _series = obj[0].series + 1
        obj.update(series=_series, desc=_key_name)
    else:
        _series = 1
        Series.objects.create(key=_key, series=1, desc=_key_name)
    return _series


@login_required
def statistic(request):
    item_map = []
    with connection.cursor() as cursor:
        sql = """select category,item_code,spec,cost_center,unitId,unitName, sum(qty)-sum(received_qty) sum_qty from inventory_appliedform a, inventory_applieditem b, users_unit c 
                 where a.form_no = b.applied_form_id and a.unit_id = c.id
                    and a.status_id=2
                    group by category,item_code,spec,cost_center,unitId,unitName having sum_qty > 0"""
        rows = django_go_sql(sql)

        item_map = sorted((list(set((dic["item_code"] for dic in rows)))))
        unit_map = (list(set((dic["unitName"] for dic in rows))))

    html_row = "<table border='1' class='table table-bordered table-striped'>"
    # 部門HEADER
    html_row += "<tr><td style='width:40px'></td><td style='width:200px'></td>"
    for unit_data in unit_map:
        html_row += "<td style='text-align:center;width:50px;'>" + unit_data + "</td>"
    html_row += "<td style='text-align:center;width:50px'>加總</td>"
    html_row += "</tr>"

    # 統計數值
    for item_data in item_map:
        item = Item.objects.get(item_code=item_data)
        sum_qty = 0
        html_row += "<tr><td>"+item.item_type.category.category_name+"</td><td>"+item.spec+"</td>"
        for unit_data in unit_map:
            value = ""
            for row in rows:
                if row["spec"]==item.spec and row["unitName"]==unit_data:
                    value = row["sum_qty"]
                    sum_qty += value
            if value:
                html_row += "<td style='text-align:right;'>"+str(value)+"</td>"
            else:
                html_row += "<td style='text-align:right;'>0</td>"
        html_row += "<td style='text-align:right;color:blue;font-weight:bold;'>"+str(sum_qty)+"</td>"
        html_row += "</tr>"
    html_row += "</table>"


    return render(request, 'inventory/statistic.html', locals())


# 回主頁
def main(request):
    form = OfficeInvForm()

    return render(request, 'inventory/application.html', locals())


@login_required
def import_excel(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('files1')
            if excel_file:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.worksheets[0]
                for iRow in range(2, sheet.max_row+1):
                    if not sheet.cell(row=iRow, column=1).value:
                        break
                    category = sheet.cell(row=iRow, column=2).value
                    category = ItemCategory.objects.get(category_name=category)
                    type = sheet.cell(row=iRow, column=3).value
                    type = ItemType.objects.get(type_name=type)
                    vendor_code = sheet.cell(row=iRow, column=4).value
                    spec = sheet.cell(row=iRow, column=5).value
                    unit = sheet.cell(row=iRow, column=6).value
                    price = sheet.cell(row=iRow, column=8).value

                    item = Item()
                    _key = type.category.catogory_code+type.type_code
                    _key_name = type.type_name
                    series = get_series_number(_key, _key_name)
                    item_code = type.category.catogory_code+type.type_code+str(series).zfill(5)
                    item = Item.objects.update_or_create(item_type=type, vendor_code=vendor_code, spec=spec, unit=unit,
                                                         price=price, create_by=request.user, update_by=request.user,
                                                             defaults={'item_code': item_code,})
        except Exception as e:
            print(e)

    return render(request, 'inventory/import.html', locals())


@login_required
def apply_list(request):
    list = AppliedForm.objects.all()
    if request.method == 'POST':
        _status = request.POST['status']
        _start_date = str(request.POST['start_date']).replace('/', '-')
        _due_date = str(request.POST['due_date']).replace('/', '-')

        if _status:
            list = list.filter(status=_status)

        if _start_date and _due_date:
            list = list.filter(apply_date__gte=_start_date, apply_date__lte=_due_date)
        form = InvAppliedHistoryForm(request.POST)
    else:
        _start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        _due_date = datetime.now().strftime('%Y-%m-%d')
        list = list.filter(apply_date__gte=_start_date, apply_date__lte=_due_date)
        form = InvAppliedHistoryForm()
    list = list.filter(Q(requester=request.user) | Q(approver=request.user))
    list = list.order_by('-apply_date')
    return render(request, 'inventory/list.html', locals())


@login_required
def approve(request):
    list = AppliedForm.objects.filter(status_id=1, approver=request.user)
    return render(request, 'inventory/approve.html', locals())


def unlock(key):
    apply_date = key[0:4] + "-" + key[4:6] + "-" + key[6:8]
    pk = key[8:]
    return apply_date, pk

@login_required
def agree(request, key):
    if request.method == 'GET':
        apply_date, pk = unlock(key)
        apply = AppliedForm.objects.get(pk=pk, apply_date=apply_date)
        apply.status = FormStatus.objects.get(pk=2)
        apply.approve_time = datetime.now()
        apply.save()
    return redirect(reverse('inv_approve'))


def mail_agree(request, key):
    if request.method == 'GET':
        apply_date, pk = unlock(key)
        form = AppliedForm.objects.get(pk=pk, apply_date=apply_date)
        items = form.applied_form_item.all().order_by('category')

        if form.status.id in [2, 6]:
            action = "done"
        else:
            form.status = FormStatus.objects.get(pk=2)
            apply.approve_time = datetime.now()
            form.save()
            action = "agree"

        return render(request, 'inventory/email_template.html', locals())


@login_required
def reject(request, key):
    if request.method == 'GET':
        apply_date, pk = unlock(key)
        apply = AppliedForm.objects.get(pk=pk, apply_date=apply_date)
        apply.status = FormStatus.objects.get(pk=6)
        apply.approve_time = datetime.now()
        apply.save()

        address = apply.requester.email
        subject = '總務用品請領單，退單通知!!!!'
        send_template_email(subject, action="reject", pk=apply.pk, address=address)

    return redirect(reverse('inv_approve'))


def mail_reject(request, key):
    if request.method == 'GET':
        apply_date, pk = unlock(key)
        form = AppliedForm.objects.get(pk=pk, apply_date=apply_date)
        items = form.applied_form_item.all().order_by('category')

        if form.status.id in [2, 6]:
            action = "done"
        else:
            form.status = FormStatus.objects.get(pk=6)
            form.approve_time = datetime.now()
            form.save()

            action = "reject"
            address = form.requester.email
            subject = '總務用品請領單，退單通知!!!!'
            send_template_email(subject, action, pk=form.pk, address=address)

        return render(request, 'inventory/email_template.html', locals())


@login_required
def delete(request, pk):
    if request.method == 'GET':
        apply = AppliedForm.objects.get(pk=pk)
        apply.status = FormStatus.objects.get(pk=4)
        apply.update_by = request.user
        apply.save()

    return redirect(reverse('inv_list'))


@login_required
def apply(request):
    if request.method == 'POST':
        hidCart_list = request.POST.get('hidCart_list')
        if hidCart_list:
            items = json.loads(hidCart_list)

        unit = request.POST.get('unit')
        requester = request.POST.get('requester')
        apply_date = request.POST.get('apply_date')
        ext_number = request.POST.get('ext_number')
        reason = request.POST.get('reason')

        try:
            apply = AppliedForm()
            YYYYMM = datetime.now().strftime("%Y%m")
            key = "OR"+YYYYMM
            apply.form_no = key + str(get_series_number(key, "文具請領")).zfill(3)
            apply.unit = Unit.objects.get(pk=unit)
            apply.requester = CustomUser.objects.get(id=requester)
            apply.apply_date = apply_date
            apply.ext_number = ext_number
            apply.reason = reason
            apply.status = FormStatus.objects.get(pk=1)
            apply.create_by = request.user
            apply.update_by = request.user
            apply.approver = request.user.manager
            apply.save()

            if request.FILES.get('file1'):
                request_file = Apply_attachment(files=request.FILES['file1'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('file2'):
                request_file = Apply_attachment(files=request.FILES['file2'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('file3'):
                request_file = Apply_attachment(files=request.FILES['file3'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()

            for item in items:
                obj = AppliedItem()
                obj.category = item['category']
                obj.item_code = item['item_code']
                obj.spec = item['spec']
                obj.qty = item['qty']
                obj.unit = item['unit']
                obj.comment = item['comment']
                obj.applied_form = apply
                obj.save()

            address = apply.requester.manager.email
            subject = '總務用品請領單簽核通知'
            send_template_email(subject, action="email", pk=apply.pk, address=address)

        except Exception as e:
            print(e)

        return redirect(apply.get_absolute_url())

    form = OfficeInvForm(initial={"unit": request.user.unit, "requester": request.user})
    form.fields["requester"].queryset = CustomUser.objects.filter(emp_no=request.user.emp_no).all()
    search_form = SearchForm()
    attach_form = AttachmentForm()
    return render(request, 'inventory/application.html', locals())


@login_required
def detail(request, pk):
    try:
        form = AppliedForm.objects.get(pk=pk)
        items = form.applied_form_item.all().order_by('category')

        status_html = get_invform_status_dropdown(form)

        if form.status.id == 1 and form.requester.manager == request.user:
            isApprover = True

        if form.requester == request.user:
            isCreater = True

        files = Apply_attachment.objects.filter(apply=form)
        for file in files:
            file.files.filename = file.files.name[file.files.name.rindex('/')+1:]
        key = "{apply_date}{series}".format(apply_date=form.apply_date.replace('-', ''),
                                           series=form.pk)


    except AppliedForm.DoesNotExist:
        raise Http404('Form does not exist')


    return render(request, 'inventory/detail.html', locals())


#API類別
def TypeAPI(request, category_id):
    type_data = ItemType.objects.filter(category_id=int(category_id)).values('id', 'type_name').order_by('type_code')
    type_list = []
    for data in type_data:
        type_list.append({'id': data['id'], 'type_name': data['type_name']})

    return JsonResponse(type_list, safe = False)


#API類別
def ItemAPI(request):
    item_list = []
    if request.method == 'POST':
        category_id = request.POST.get('category_id')
        type_id = request.POST.get('type_id')
        keyword = request.POST.get('keyword')

        if type_id:
            item_data = Item.objects.all().values('item_code', 'spec', 'unit')
        else:
            item_data = Item.objects.exclude(spec__contains="自行輸入").values('item_code', 'spec', 'unit')

        if category_id:
            item_type = ItemType.objects.filter(category_id=category_id)
            item_data = item_data.filter(item_type__in=item_type).values('item_code', 'spec', 'unit')
        if type_id:
            item_data = item_data.filter(item_type_id=int(type_id)).values('item_code', 'spec', 'unit')
        if keyword:
            query = Q(spec__icontains=keyword)
            item_data = item_data.filter(query).values('item_code', 'spec', 'price', 'unit')

        for data in item_data:
            category = ItemCategory.objects.get(id=category_id)
            item_list.append({'item_code': data['item_code'], 'spec': data['spec'], 'unit': data['unit'], 'category': category.category_name})

    return JsonResponse(item_list, safe=False)


def change_status(request):
    if request.POST:
        form_id = request.POST.get('form_id')
        status_id = request.POST.get('status_id')

        status = FormStatus.objects.get(pk=status_id)
        obj = AppliedForm.objects.get(pk=form_id)
        obj.status = status
        obj.save()

        return redirect(obj.get_absolute_url())


def recieved(request, pk):
    item = AppliedItem.objects.get(pk=pk)
    if request.POST:
        applied_form = AppliedForm.objects.get(pk=item.applied_form.pk)
        form = OfficeItemForm(request.POST, instance=item)
        form.save()
        return redirect(applied_form.get_absolute_url())

    form = OfficeItemForm(instance=item)
    return render(request, 'inventory/recieved.html', locals())


def pr_apply(request):
    if request.method == 'POST':
        hidCart_list = request.POST.get('hidCart_list')
        if hidCart_list:
            items = json.loads(hidCart_list)

        unit = request.POST.get('unit')
        requester = request.POST.get('requester')
        apply_date = request.POST.get('apply_date')
        ext_number = request.POST.get('ext_number')
        reason = request.POST.get('reason')

        try:
            apply = AppliedForm()
            apply.unit = Unit.objects.get(pk=unit)
            apply.requester = CustomUser.objects.get(id=requester)
            apply.apply_date = apply_date
            apply.ext_number = ext_number
            apply.reason = reason
            apply.status = FormStatus.objects.get(pk=1)
            apply.create_by = request.user
            apply.save()

            if request.FILES.get('file1'):
                request_file = Apply_attachment(files=request.FILES['file1'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('file2'):
                request_file = Apply_attachment(files=request.FILES['file2'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()
            if request.FILES.get('file3'):
                request_file = Apply_attachment(files=request.FILES['file3'])
                request_file.apply = apply
                request_file.create_by = request.user
                request_file.save()

            for item in items:
                obj = AppliedItem()
                obj.item_code = item['item_code']
                obj.spec = item['spec']
                obj.qty = item['qty']
                obj.unit = item['unit']
                obj.applied_form = apply
                obj.save()

            action = "email"
            email = apply.requester.manager.email
            if email:
                # 電子郵件內容樣板
                pk = apply.pk
                form = AppliedForm.objects.get(pk=pk)
                files = Apply_attachment.objects.filter(apply=form)
                for file in files:
                    file.files.filename = file.files.name[file.files.name.rindex('/') + 1:]
                email_template = render_to_string('inventory/email_template.html', locals())

                email = EmailMessage(
                    '總務用品請領單簽核通知',  # 電子郵件標題
                    email_template,  # 電子郵件內容
                    settings.EMAIL_HOST_USER,  # 寄件者
                    [email]  # 收件者
                )
                email.fail_silently = False
                email.content_subtype = 'html'
                email.send()
                print("郵件成功寄出")

        except Exception as e:
            print(e)

        return redirect(apply.get_absolute_url())

    form = OfficeInvForm()
    search_form = SearchForm()
    attach_form = AttachmentForm()
    return render(request, 'inventory/pr_apply.html', locals())


def search(request):
    search_form = SearchForm()
    return render(request, 'inventory/search.html', locals())
